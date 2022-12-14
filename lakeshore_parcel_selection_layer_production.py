#-------------------------------------------------------------------------------
# Name:        lakeshore_parcel_selection_layer_production.py
# Purpose:      This script iterates over an input XLSX file to find Lake ID numbers,
#                   selects those lakes from the state lake layer, uses those
#                   selected lakes to select parcels from the state parcel layer, and
#                   then produces a feature class with the selected parcels.
#
# Author:      draleigh
#
# Created:     11 November 2022
# Copyright:   (c) draleigh 2022
# Licence:     <your licence>
#-------------------------------------------------------------------------------
#



# Import modules and packages
import os,arcpy,time,datetime,copy,openpyxl                                                 # list of modules that will be needed
from datetime import date, timedelta                                                        # extract the sub modules
from openpyxl import load_workbook                                                          # extract the sub module
from arcpy import env                                                                       # extract the sub module
from arcpy.sa import *                                                                      # import the Spatial Analyst toolbar
arcpy.CheckOutExtension("Spatial")                                                          # check out the Spatial Analyst toolbar

# Date Management
start_time = time.time()                                                                    # define the start time of the script
today = date.today()                                                                        # define the date of the script's run
today_year = today.year                                                                     # define the year of the script's run
today_month = today.month                                                                   # define the month of the script's run
today_day = today.day                                                                       # define the day of the script's run
#print(today_day)                                                                           # print the day (if desired)
today_year_str = today.strftime("%Y")                                                       # produce the string value of the script's run date year
today_month_str = today.strftime("%m")                                                      # produce the string value of the script's run date month
today_day_str = today.strftime("%d")                                                        # produce the string value of the script's run date day
today_full_str = today_year_str + today_month_str + today_day_str                           # set up string value of the script's run date (e.g. for January 5, 2002: 20020105
#print(today_full_str)                                                                      # print out today_full_str in the terminal


# Define Data Locations and Other Variables
Lakeshore_GDB = r"*folderpath*\Lakeshore_Parcel_Ownership_Analysis.gdb"                         # define the location of the geodatabase to which you'll send the products of the script
Lakes = r"*folderpath*\dnr_hydro_features_all"                                                   # define the location of the feature class that you'll pull the selecting features from
Parcels = r"*folderpath*\plan_parcels_minnesota"                                                # define the location of the feature class that you'll ultimately select parcels from
Input = r"*folderpath*\Test_Lakes.xlsx"                                                         # define the location of the file that holds the lake ID values
Parcel_Selection = os.path.join(Lakeshore_GDB, 'Parcel_Selection_' + today_full_str)            # define the file path of the feature class that will be produced


# Obtain Lake IDs by the column specified (but need to know the number, not name of column)
Lake_IDs_full = []                                                                              # set up an empty list to hold values
wb = openpyxl.load_workbook(Input)                                                              # use the openpyxl module to load the workbook that you've already defined
sheet = wb['Test_Lakes']                                                                        # define the sheet according to the name of the sheet in the workbook
for val in sheet.iter_rows(min_col=2):                                                          # set up for loop to iterate over rows in the sheet for the second column (this one holds the lake ID values)
    Lake_IDs_full.append(val[0].value)                                                          # add the initial value from that row (limited to start at column 2) to the previously set-up list
    print(val[0].value)                                                                         # display the value to the terminal (not necessary)
Lake_IDs = Lake_IDs_full[1:]                                                                    # remove the first value in the list (the header of the column, if necessary)
print(Lake_IDs)                                                                                 # print the entire list (if necessary)

# Use the Lake IDs (Lake_IDs list) to select the lakes from the GDRS Hydrography shapefile
arcpy.MakeFeatureLayer_management(Lakes, 'Lakes')                                               # set up a temporary feature layer for use with the arcpy module
Lake_String = ''                                                                                # set up an empty string
for x in range(len(Lake_IDs)):                                                                  # start a for loop to iterate over the list that has been previously defined
    Lake_String = Lake_String + "DOWLKNUM = '" + str(Lake_IDs[x]) + "' OR "                     # the for loop will add values to the string for each iteration to produce an expression in SQL
Lake_Expression = Lake_String[:-4]                                                              # once the for loop is finished, define a new string that takes off the final 4 digits (which are " OR ")
print(Lake_Expression)                                                                          # print the entire string (if necessary)

arcpy.SelectLayerByAttribute_management('Lakes', "NEW_SELECTION", Lake_Expression)              # utilize the 'SelectLayerByAttribute_management' function to select from the temporary 'Lakes' layer with a "NEW SELECTION" utilizing the previously-defined string expression
arcpy.MakeFeatureLayer_management(Parcels, 'Parcels')                                           # set up a temporary feature layer for use with the arcpy module
arcpy.SelectLayerByLocation_management('Parcels', "INTERSECT", 'Lakes', "", "NEW_SELECTION")    # using the selected features within the 'Lakes' layer, select all 'Parcels' polygons that "INTERSECT" with the 'Lakes' polygons
arcpy.CopyFeatures_management('Parcels', Parcel_Selection)                                      # copy the selected features from 'Parcels' to make a new feature class whose file path has been previously defined


